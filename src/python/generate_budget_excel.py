#!/usr/bin/env python3
"""Fill the official ABC budget template with cart items from the web app."""

import json
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

COL_ITEM_NO = 1
COL_DESC = 2
COL_QTY = 3
COL_UNIT = 4
COL_MARKET = 5
COL_TOTAL = 11
COL_UNIT_COST = 12
COL_LAST = 12
DEFAULT_SHEET = "ABC (2)"
MONEY_FORMAT = "#,##0.00"

ALIGN_ITEM_NO = Alignment(horizontal="center", vertical="center")
ALIGN_DESC = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_QTY = Alignment(horizontal="center", vertical="center")
ALIGN_UNIT = Alignment(horizontal="center", vertical="center")
ALIGN_MONEY = Alignment(horizontal="right", vertical="center")
ALIGN_CATEGORY = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_TOTAL_LABEL = Alignment(horizontal="right", vertical="center")


def parse_revised_date(value: str) -> str:
    if not value:
        return datetime.now().strftime("%B %d, %Y")
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value.strip(), fmt).strftime("%B %d, %Y")
        except ValueError:
            continue
    return value.strip()


def group_items_by_category(items: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = {}
    for entry in items:
        category = (entry.get("category") or "Uncategorized").strip()
        grouped.setdefault(category, []).append(entry)
    return dict(sorted(grouped.items(), key=lambda pair: pair[0].lower()))


def copy_cell_style(source, target) -> None:
    if not source.has_style:
        return
    target.font = copy(source.font)
    target.border = copy(source.border)
    target.fill = copy(source.fill)
    target.protection = copy(source.protection)


def copy_row_style(ws, source_row: int, target_row: int, max_col: int = COL_LAST) -> None:
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))


def find_row_containing(ws, needle: str, start_row: int = 1) -> int | None:
    needle_lower = needle.lower()
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value is not None and needle_lower in str(value).lower():
                return row
    return None


def find_data_bounds(ws) -> tuple[int, int, int, int]:
    header_row = find_row_containing(ws, "ITEM NO.")
    if header_row is None:
        raise ValueError("Could not locate the item table header in the template.")

    data_start = None
    category_style_row = None
    item_style_row = None

    for row in range(header_row + 1, ws.max_row + 1):
        desc = ws.cell(row, COL_DESC).value
        if desc is None:
            continue
        text = str(desc).strip()
        if not text or text.startswith("("):
            continue
        if data_start is None:
            data_start = row
        if ws.cell(row, COL_ITEM_NO).value in (None, ""):
            if category_style_row is None:
                category_style_row = row
        elif item_style_row is None:
            item_style_row = row

    if data_start is None or category_style_row is None or item_style_row is None:
        raise ValueError("Could not locate data rows in the template.")

    total_row = None
    for row in range(data_start, ws.max_row + 1):
        for col in (10, COL_TOTAL):
            value = ws.cell(row, col).value
            if value is not None and str(value).strip().upper() == "TOTAL":
                total_row = row
                break
        if total_row is not None:
            break

    if total_row is None:
        raise ValueError("Could not locate the TOTAL row in the template.")

    return data_start, total_row, category_style_row, item_style_row


def find_signature_row(ws, after_row: int) -> int | None:
    return find_row_containing(ws, "PREPARED AND SUBMITTED BY", start_row=after_row)


def update_header_fields(ws, department: str, project_title: str, revised_date: str) -> None:
    department = department.upper()
    ws.cell(2, COL_DESC, department)
    ws.cell(5, 1, f"Revised on: {revised_date}")
    ws.cell(7, 1, project_title)


def apply_category_row(ws, row: int, category: str, style_row: int) -> None:
    copy_row_style(ws, style_row, row)
    ws.row_dimensions[row].height = ws.row_dimensions[style_row].height or 24

    for col in range(1, COL_LAST + 1):
        cell = ws.cell(row, col)
        cell.value = None
        if col == COL_DESC:
            cell.value = category.upper()
            cell.alignment = ALIGN_CATEGORY
        else:
            cell.alignment = ALIGN_CATEGORY


def apply_item_row(ws, row: int, item_no: int, entry: dict, style_row: int) -> float:
    copy_row_style(ws, style_row, row)
    ws.row_dimensions[row].height = ws.row_dimensions[style_row].height or 22

    qty = float(entry.get("quantity") or 0)
    unit_cost = float(entry.get("unit_cost") or 0)
    total_cost = round(qty * unit_cost, 2)

    values = {
        COL_ITEM_NO: item_no,
        COL_DESC: entry.get("item_name") or "",
        COL_QTY: qty,
        COL_UNIT: entry.get("unit") or "",
        COL_MARKET: unit_cost if unit_cost else None,
        COL_TOTAL: total_cost,
        COL_UNIT_COST: unit_cost if unit_cost else None,
    }

    for col in range(1, COL_LAST + 1):
        cell = ws.cell(row, col)
        cell.value = values.get(col)
        if col in (COL_MARKET, COL_TOTAL, COL_UNIT_COST) and cell.value is not None:
            cell.number_format = MONEY_FORMAT

    ws.cell(row, COL_ITEM_NO).alignment = ALIGN_ITEM_NO
    ws.cell(row, COL_DESC).alignment = ALIGN_DESC
    ws.cell(row, COL_QTY).alignment = ALIGN_QTY
    ws.cell(row, COL_UNIT).alignment = ALIGN_UNIT
    for col in (COL_MARKET, COL_TOTAL, COL_UNIT_COST):
        ws.cell(row, col).alignment = ALIGN_MONEY

    for col in range(6, 10):
        ws.cell(row, col, None)

    return total_cost


def apply_total_row(ws, row: int, grand_total: float, style_row: int) -> None:
    copy_row_style(ws, style_row, row)
    ws.row_dimensions[row].height = ws.row_dimensions[style_row].height or 22

    for col in range(1, COL_LAST + 1):
        ws.cell(row, col, None)

    ws.cell(row, 10, "TOTAL").alignment = ALIGN_TOTAL_LABEL
    total_cell = ws.cell(row, COL_TOTAL, round(grand_total, 2))
    total_cell.number_format = MONEY_FORMAT
    total_cell.alignment = ALIGN_MONEY


def remove_blank_rows(ws, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    ws.delete_rows(start_row, end_row - start_row + 1)


def finalize_sheet(ws, last_row: int) -> None:
    ws.sheet_view.view = "normal"
    ws.row_breaks.brk = []
    ws.col_breaks.brk = []
    ws.print_area = f"A1:{get_column_letter(COL_LAST)}{max(last_row + 2, 30)}"


def generate_from_template(payload: dict, template_path: Path, output_path: Path) -> None:
    department = (payload.get("department") or "TOURISM").strip()
    project_title = (payload.get("project_title") or "").strip()
    revised_date = parse_revised_date(payload.get("revised_date", ""))
    items = payload.get("items") or []
    sheet_name = payload.get("sheet_name") or DEFAULT_SHEET

    if not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' was not found in the template workbook.")
    ws = wb[sheet_name]

    data_start, total_row, category_style_row, item_style_row = find_data_bounds(ws)
    signature_row = find_signature_row(ws, total_row) or (total_row + 4)
    total_style_row = total_row

    update_header_fields(ws, department, project_title, revised_date)

    grouped = group_items_by_category(items)
    new_row_count = sum(1 + len(category_items) for category_items in grouped.values()) + 1
    old_row_count = total_row - data_start + 1

    trailing_blank_end = signature_row - 1 if signature_row > total_row else total_row
    delete_through = max(total_row, trailing_blank_end)
    rows_to_delete = delete_through - data_start + 1

    ws.delete_rows(data_start, rows_to_delete)

    signature_row = find_signature_row(ws, data_start) or (data_start + new_row_count + 2)
    ws.insert_rows(data_start, new_row_count)

    current_row = data_start
    grand_total = 0.0

    for category, category_items in grouped.items():
        apply_category_row(ws, current_row, category, category_style_row)
        current_row += 1

        item_no = 0
        for entry in sorted(category_items, key=lambda x: (x.get("item_name") or "").lower()):
            item_no += 1
            grand_total += apply_item_row(ws, current_row, item_no, entry, item_style_row)
            current_row += 1

    apply_total_row(ws, current_row, grand_total, total_style_row)
    total_row = current_row

    signature_row = find_signature_row(ws, total_row + 1)
    if signature_row and signature_row > total_row + 1:
        remove_blank_rows(ws, total_row + 1, signature_row - 1)

    finalize_sheet(ws, total_row + 8)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def resolve_template_path(payload: dict) -> Path:
    if payload.get("template_path"):
        return Path(payload["template_path"])
    base = Path(__file__).resolve().parent.parent.parent
    candidates = [
        base / "templates" / "ABC_TOURISM_NATIONAL_ARTS_MONTH.xlsx",
        base / "excel_files" / "ABC TOURISM NATIONAL ARTS MONTH.xlsx",
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return candidates[0]


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: generate_budget_excel.py <input.json> <output.xlsx>", file=sys.stderr)
        return 1

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    payload = json.loads(input_path.read_text(encoding="utf-8-sig"))
    template_path = resolve_template_path(payload)

    generate_from_template(payload, template_path, output_path)
    print(f"Generated {output_path} from template {template_path.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
